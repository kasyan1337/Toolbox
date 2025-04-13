import os
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import io

GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")


def sanitize_value(val):
    if isinstance(val, str):
        return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", val)
    return val


def read_excel_file(file_path):
    return pd.read_excel(file_path, dtype=str)


def compute_excel_diff(df_input, df_output):
    # Get union of indices and determine columns for alignment
    all_index = sorted(set(df_input.index).union(set(df_output.index)))
    all_columns = (
        list(df_input.columns) if not df_input.empty else list(df_output.columns)
    )
    df_input = df_input.reindex(index=all_index, columns=all_columns)
    df_output = df_output.reindex(index=all_index, columns=all_columns)
    # Create masks where values differ and the differing cell in one of the DataFrames is not empty
    addition_mask = (df_input.fillna("") != df_output.fillna("")) & df_output.fillna(
        ""
    ).ne("")
    missing_mask = (df_input.fillna("") != df_output.fillna("")) & df_input.fillna(
        ""
    ).ne("")
    return df_input, df_output, addition_mask, missing_mask


def apply_diff_highlights(excel_path, diff_mask, output_path, fill):
    wb = load_workbook(excel_path)
    ws = wb.active
    # Iterate over diff mask rows; row numbers for Excel are offset by 2 (header at row 1)
    for i in range(len(diff_mask)):
        if diff_mask.iloc[i].any():
            excel_row = i + 2
            for cell in ws[excel_row]:
                cell.fill = fill
    wb.save(output_path)


def document_diff(
    df_input,
    df_output,
    addition_mask,
    missing_mask,
    file_name,
    input_file,
    output_file,
    summary_file,
):
    total_additions = addition_mask.values.sum()
    total_missing = missing_mask.values.sum()
    lines = []
    lines.append(f"File: {file_name}")
    lines.append(f"Input Excel: {input_file}")
    lines.append(f"Output Excel: {output_file}")
    lines.append(f"Total additions (green): {total_additions}")
    lines.append(f"Total missing (red): {total_missing}")
    lines.append("Detailed differences per cell:")
    # List each cell where the output contains a new (added) value
    for idx in addition_mask.index:
        for col in addition_mask.columns:
            if addition_mask.at[idx, col]:
                lines.append(
                    f"  Row {int(idx)+2}, Column {col}: input='{df_input.at[idx, col]}' output='{df_output.at[idx, col]}' (ADDED)"
                )
    # List each cell where the output is missing a value present in the input
    for idx in missing_mask.index:
        for col in missing_mask.columns:
            if missing_mask.at[idx, col]:
                lines.append(
                    f"  Row {int(idx)+2}, Column {col}: input='{df_input.at[idx, col]}' output='{df_output.at[idx, col]}' (MISSING)"
                )
    with open(summary_file, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def process_excel_pair(input_path, output_path, diff_folder):
    df_input = read_excel_file(input_path)
    df_output = read_excel_file(output_path)
    # Ensure matching columns exist even if one file is empty
    if df_input.empty and not df_output.empty:
        df_input = pd.DataFrame(columns=df_output.columns)
    elif df_output.empty and not df_input.empty:
        df_output = pd.DataFrame(columns=df_input.columns)
    df_input = df_input.reset_index(drop=True)
    df_output = df_output.reset_index(drop=True)
    df_input, df_output, addition_mask, missing_mask = compute_excel_diff(
        df_input, df_output
    )
    diff_total = addition_mask.values.sum() + missing_mask.values.sum()
    file_name = os.path.basename(input_path)
    if diff_total == 0:
        print(f"No differences detected in {file_name}; skipping output.")
        return
    # Create subdirectories for output highlights and the summary
    input_diff_path = os.path.join(diff_folder, "input_diff", file_name)
    output_diff_path = os.path.join(diff_folder, "output_diff", file_name)
    summary_dir = os.path.join(diff_folder, "differences_summary")
    os.makedirs(os.path.join(diff_folder, "input_diff"), exist_ok=True)
    os.makedirs(os.path.join(diff_folder, "output_diff"), exist_ok=True)
    os.makedirs(summary_dir, exist_ok=True)
    # Apply highlights: red for missing in output, green for added in output
    apply_diff_highlights(input_path, missing_mask, input_diff_path, RED_FILL)
    apply_diff_highlights(output_path, addition_mask, output_diff_path, GREEN_FILL)
    summary_file = os.path.join(
        summary_dir, f"{os.path.splitext(file_name)[0]}_summary.txt"
    )
    document_diff(
        df_input,
        df_output,
        addition_mask,
        missing_mask,
        file_name,
        input_path,
        output_path,
        summary_file,
    )
    print(f"Processed {file_name}: summary saved at {summary_file}")


def process_diff(input_folder, output_folder, diff_folder):
    # Verify both input and output folders exist
    if not (os.path.exists(input_folder) and os.path.exists(output_folder)):
        print("Input or output folder does not exist.")
        return
    os.makedirs(diff_folder, exist_ok=True)
    # Get list of Excel files in the input folder
    input_files = [f for f in os.listdir(input_folder) if f.lower().endswith(".xlsx")]
    # For each input file, find the corresponding file in the output folder (case-insensitive match)
    for file in input_files:
        matching_file = None
        for out_file in os.listdir(output_folder):
            if out_file.lower() == file.lower():
                matching_file = out_file
                break
        if matching_file:
            input_path = os.path.join(input_folder, file)
            output_path = os.path.join(output_folder, matching_file)
            process_excel_pair(input_path, output_path, diff_folder)
        else:
            print(f"No matching output file for {file}")


def main():
    # Use current working directory as the base folder
    base_dir = os.getcwd()
    # Define the folder paths: 'input' for the original Excel files and 'output' for the changed ones
    input_folder = os.path.join(base_dir, "input")
    output_folder = os.path.join(base_dir, "output")
    # Differences (highlights and summaries) are stored in the 'differences' folder
    diff_folder = os.path.join(base_dir, "differences")
    process_diff(input_folder, output_folder, diff_folder)


if __name__ == "__main__":
    main()
